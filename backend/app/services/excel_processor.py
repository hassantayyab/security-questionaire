"""
Excel processing service using openpyxl for questionnaire data extraction
"""

import io
from typing import List, Dict, Any
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import logging

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Excel processing service using openpyxl library"""
    
    def extract_questions_from_bytes(self, excel_bytes: bytes) -> List[Dict[str, Any]]:
        """
        Extract questions from Excel file bytes
        
        Expected format:
        - Column A: Question text
        - Column B: Answer (optional, can be empty)
        - First row can be headers (will be skipped if detected)
        
        Args:
            excel_bytes: Excel file content as bytes
            
        Returns:
            List[Dict]: List of questions with text and optional answers
            
        Raises:
            Exception: If Excel processing fails
        """
        try:
            # Read Excel content into BytesIO object
            excel_stream = io.BytesIO(excel_bytes)
            
            # Load workbook from memory
            workbook = openpyxl.load_workbook(excel_stream, read_only=True)
            
            # Get the first worksheet
            worksheet = workbook.active
            
            logger.info(f"Processing Excel worksheet: {worksheet.title}")
            
            questions = []
            row_count = 0
            
            # Iterate through rows
            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), 1):
                # Skip empty rows
                if not any(row):
                    continue
                
                # Get question text from first column
                question_text = str(row[0]).strip() if row[0] else ""
                
                # Skip if no question text
                if not question_text:
                    continue
                
                # Check if this looks like a header row (skip it)
                if row_idx == 1 and self._is_header_row(question_text):
                    logger.info("Skipping header row")
                    continue
                
                # Get answer from second column if present
                answer = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                
                question_data = {
                    "question_text": question_text,
                    "answer": answer if answer else None,
                    "status": "unapproved",
                    "row_number": row_idx
                }
                
                questions.append(question_data)
                row_count += 1
                
                logger.info(f"Extracted question {row_count}: {question_text[:50]}...")
            
            workbook.close()
            
            if not questions:
                raise Exception("No valid questions found in Excel file")
            
            logger.info(f"Successfully extracted {len(questions)} questions from Excel file")
            
            return questions
            
        except Exception as e:
            logger.error(f"Error processing Excel file: {str(e)}")
            raise Exception(f"Error processing Excel file: {str(e)}")
    
    def _is_header_row(self, text: str) -> bool:
        """
        Check if a row appears to be a header row
        
        Args:
            text: Text from first column of row
            
        Returns:
            bool: True if likely a header row
        """
        header_indicators = [
            "question", "questions", "query", "queries",
            "item", "items", "description", "text"
        ]
        
        return any(indicator in text.lower() for indicator in header_indicators)
    
    def extract_questions_from_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Extract questions from Excel file path (alternative method)
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List[Dict]: List of questions
        """
        try:
            with open(file_path, 'rb') as file:
                excel_bytes = file.read()
                return self.extract_questions_from_bytes(excel_bytes)
        except FileNotFoundError:
            raise Exception(f"Excel file not found: {file_path}")
        except Exception as e:
            raise Exception(f"Error reading Excel file: {str(e)}")
    
    def create_export_excel(self, questions: List[Dict[str, Any]], filename: str = "questionnaire_export.xlsx") -> bytes:
        """
        Create Excel file from questions and answers for export
        
        Args:
            questions: List of question dictionaries with answers
            filename: Name for the exported file
            
        Returns:
            bytes: Excel file content as bytes
        """
        try:
            # Create new workbook
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Questionnaire Answers"
            
            # Add headers
            headers = ["Question", "Answer", "Status"]
            for col_idx, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col_idx, value=header)
            
            # Add questions and answers
            for row_idx, question in enumerate(questions, 2):
                worksheet.cell(row=row_idx, column=1, value=question.get("question_text", ""))
                worksheet.cell(row=row_idx, column=2, value=question.get("answer", ""))
                worksheet.cell(row=row_idx, column=3, value=question.get("status", ""))
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 100)  # Cap at 100 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            excel_stream = io.BytesIO()
            workbook.save(excel_stream)
            excel_stream.seek(0)
            
            workbook.close()
            
            logger.info(f"Created Excel export with {len(questions)} questions")
            
            return excel_stream.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating Excel export: {str(e)}")
            raise Exception(f"Error creating Excel export: {str(e)}")
    
    def get_excel_info(self, excel_bytes: bytes) -> Dict[str, Any]:
        """
        Get information about Excel file
        
        Args:
            excel_bytes: Excel file content as bytes
            
        Returns:
            Dict: Information about the Excel file
        """
        try:
            excel_stream = io.BytesIO(excel_bytes)
            workbook = openpyxl.load_workbook(excel_stream, read_only=True)
            
            worksheet = workbook.active
            
            # Count rows with content
            row_count = 0
            for row in worksheet.iter_rows(values_only=True):
                if any(row):
                    row_count += 1
            
            info = {
                "worksheet_name": worksheet.title,
                "total_rows": row_count,
                "max_column": worksheet.max_column,
                "max_row": worksheet.max_row
            }
            
            workbook.close()
            
            return info
            
        except Exception as e:
            logger.error(f"Error getting Excel info: {str(e)}")
            return {"error": str(e)}
